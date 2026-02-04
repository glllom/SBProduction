import os
import sys
import django
from openpyxl import Workbook
from openpyxl.styles import Alignment

# Добавляем корень проекта в путь поиска модулей
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'SBProduction.settings')
django.setup()

from openpyxl import load_workbook
from production.models import *
# def load_order(order_num):
#     current_order = DBOrder.objects.get(order_number=order_num)
#     print(current_order)
#     return None

def load_products():
    wb = load_workbook('sku_3.xlsx', data_only=True)
    sheets = wb.sheetnames
    types_sheet = wb[sheets[0]]
    families_sheet = wb[sheets[1]]
    products_sheet = wb[sheets[2]]
    materials_sheet = wb[sheets[3]]

    # for row in types_sheet.iter_rows(min_row=2):
        # print(row[1].value, row[2].value)
        # DBProductType.objects.create(name=row[1].value, sku=row[2].value)

    # for row in families_sheet.iter_rows(min_row=2):
    #     DBProductFamily.objects.create(name=row[1].value, sku=row[2].value, product_type=DBProductType.objects.get(id=row[3].value))

    for row in products_sheet.iter_rows(min_row=2):
        DBProductModel.objects.create(
            product_family=DBProductFamily.objects.get(sku=row[4].value),
            description=row[1].value,
            product_type=DBProductType.objects.get(sku=row[5].value),
            sku=row[2].value,
            global_type=row[3].value)

    # for row in materials_sheet.iter_rows(min_row=2):
    #     DBComponent.objects.create(
    #         name = row[0].value,
    #         width=row[1].value,
    #         length=row[2].value,
    #         thickness=row[3].value,
    #         sku=row[4].value,
    #         group=row[5].value,
    #         component_type=row[6].value,
    #         color=row[7].value,
    #         global_type=row[8].value)


bom_dict = {
    'msh_aqua1.8x122x244': {'models': [3, 4,15,16,27,33,41,44,47,52,55,58,61], 'tag': "cover_material", 'qty': 2},
    'msh_hdf4x122x244': {'models': [7,8,11,12,13,14,23,24,25,26,48,], 'tag': "cover_material", 'qty': 2},
    'msh_hdf6x122x244': {'models': [7,8,9,10,19,20,21,22,28,29,31,32,34,35,38,39,42,43,45,46,50,51,53,54,56,57,59,60,62,63,65], 'tag': "cover_material", 'qty': 2},
    'msh_flex34': {'models': [5,6,11,12,13,14,17,17,23,24,25,26,48,49], 'tag': "filling_material", 'qty': 1.2},
    'msh_flex38': {'models': [3,4,47,52,55,58,61,], 'tag': "filling_material", 'qty': 1.2},
    'msh_flex44': {'models': [x for x in range(27, 47)], 'tag': "filling_material", 'qty': 1.2},
    'msh_loc_white': {'models': [1,2,36,40], 'tag': "cover_material", 'qty': 2},
    'msh_polystyrene4': {'models': [1,2,36,40], 'tag': "filling_material", 'qty': 2},
    'mln_pine34': {'models': [5,6,11,12,13,14,17,17,23,24,25,26,48,49], 'tag': "filling_material", 'qty': 10},
    'mln_pine38': {'models': [5,6,11,12,13,14,17,17,23,24,25,26,48,49], 'tag': "filling_material", 'qty': 10},
    'mln_fs10220w': {'models': [x for x in range(4, 15, 2)], 'tag': "frame", 'qty': 5},
    'mln_ff10207w': {'models': [x for x in range(16, 27, 2)], 'tag': "frame", 'qty': 5},
    'mln_fa10220w': {'models': [x for x in range(41, 47)], 'tag': "frame", 'qty': 5},
    'mln_fcs10220w': {'models': [27,28,29,30,31,32,64,65], 'tag': "frame", 'qty': 5},
    'mln_fws6600w': {'models': [x for x in range(33,41)], 'tag': "frame", 'qty': 1},
    'mln_fen8600b': {'models': [1, 2], 'tag': "frame", 'qty': 1},
}



# all_models = DBProductModel.objects.all()
# for model in all_models:
#     for material_sku, bom_data in bom_dict.items():
#         # print(model.sku.split('_')[1])
#         if int(model.sku.split('_')[1]) in bom_data['models']:
#             # print(material_sku, model.sku)
#             DBBomComponent.objects.create(DBComponent=DBComponent.objects.get(sku=material_sku),
#                                        product_model=model,
#                                        qty=bom_data['qty'],
#                                        tag=DBBomTag.objects.get(tag=bom_data['tag']))
#
#

bom_to_export = {}
for model in DBProductModel.objects.all():
    bom_to_export[model.sku] = [(x.DBComponent.sku, x.qty, x.tag.tag) for x in DBBomComponent.get_components_for_product(model)]

# for sku, components in bom_to_export.items():
#     print('*'*10, sku,'*'*10)
#     for component in components:
#         print('- ',component)


def export_dict_to_xlsx(data, file_name="output.xlsx"):
    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM Data"

    # Define headers
    ws.append(["Model", "Material", "Qty", "Tag"])

    current_row = 2  # Starting from the second row because of headers

    for key, tuples_list in bom_to_export.items():
        start_row = current_row

        # Write tuples to columns B, C, D
        for t in tuples_list:
            ws.cell(row=current_row, column=2).value = t[0]
            ws.cell(row=current_row, column=3).value = t[1]
            ws.cell(row=current_row, column=4).value = t[2]
            current_row += 1

        # Write the key to the first column of the first row in the group
        key_cell = ws.cell(row=start_row, column=1)
        key_cell.value = key

        # Merge column A if there is more than one tuple
        if len(tuples_list) > 1:
            # Merging from start_row to the last written row
            ws.merge_cells(
                start_row=start_row,
                start_column=1,
                end_row=current_row - 1,
                end_column=1
            )

        # Center the text in the merged cell (vertical and horizontal)
        key_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Save the file
    wb.save(file_name)
    print(f"File {file_name} saved successfully.")


# Example usage:
my_data = {
    "Order_101": [("Material A", 10, "kg"), ("Material B", 5, "m")],
    "Order_102": [("Material C", 1, "pcs")],
}

export_dict_to_xlsx(bom_to_export)